#!/usr/bin/python
import os
import sys
import argparse
import urllib3
from html.parser import HTMLParser
from openpyxl import load_workbook
from re import split

VERBOSE_DEBUG = False
PLAINTEXT = False
FILENAME = ""
WORD_SEPARATORS = [" ", "-", "—", "/", "[", "]"]
# The list below is unused due to issues with malformed output
CHARACTERS_TO_STRIP = "\r\n'‘’ "

###     TO DO   ###
# - Mark output and note whether it follows British or American English transcription
# - Add a list to the script containing all irregular verbs for easy lookup: https://myefe.com/english-irregular-verbs
# - "cannabis" returns "ˈkænəbɪsiz", "checkers"	returns "ˈtʃekəziz". This might be an issue with all verbs ending with "s"
# 
###      END    ###

# The URL which is used to look up words
baseURL = "https://www.oxfordlearnersdictionaries.com/search/english/?q="

# List of fricatives, needed to form possessive transcriptions. Source: https://pronuncian.com/introduction-to-fricatives
# Source 2 (not used): https://english.stackexchange.com/questions/5913/what-is-the-pronunciation-of-the-possessive-words-that-already-end-in-s
FRICATIVE_SOUNDS = ['v', 'f', 'ð', 'θ', 'z', 's', 'ʒ', 'ʃ', 'h']

# List of voiced consonants. Source: https://www.thoughtco.com/voiced-and-voiceless-consonants-1212092
VOICED_CONSONANTS = ['b', 'd', 'g', 'j', 'l', 'm', 'n', 'ng', 'r', 'sz', 'th', 'v', 'w', 'y', 'z']

# List of voiceless consonants. Source: https://www.thoughtco.com/voiced-and-voiceless-consonants-1212092
# Voiceless consonats which could also be found in the list of fricative sounds are discarded from the following list.
# VOICELESS_CONSONANTS = ['ch', 'f', 'k', 'p', 's', 'sh', 't', 'th']
VOICELESS_CONSONANTS = ['ch', 'f', 'k', 'p', 'sh', 't', 'th']

# List of vowels. Source: https://simple.wikipedia.org/wiki/Vowel
# Y is part of this list because the checks performed in this script only check the last letter which, if 'y', is interpreted as a vowel according to the source cited above.
VOWELS = ['a', 'а', 'e', 'i', 'o', 'u', 'y']

# A list of irregular verbs that couldn't be found at the Oxford Learner's Dictionaries.
# Source: https://www.englishpage.com/irregularverbs/irregularverbs.html#
# Also some present forms of verbs are listed below due to them being represented in a different way in the URL of OLD, e.g. sow1.
# NB! "inbreed" has been omitted
# NB! wound has two forms (both not directly found using this script) - noun and verb. The wordform has thus not been included.
# NB! Some verbs, such as "withdraw", have alternative spellings. Only one of these has been taken into consideration
IRREGULAR_VERBS = {
    "arisen": "əˈrɪzn",
    "beaten": "ˈbiːtn",
    "bid": "bɪd",
    "bidden": "ˈbɪdn",
    "bled": "bled",
    "daydreamt": "ˈdeɪdremt",
    "disproven": "ˌdɪsˈpruːvn",
    "do": "duː",
    "dreamt": "dremt",
    "dwelled": "dwelt",
    "eaten": "ˈiːtn",
    "forewent": "fɔːˈwent",
    "foregone": "fɔːˈɡɒn",
    "foresaw": "fɔːˈsɔː",
    "foreseen": "fɔːˈsiːn",
    "forgiven": "fəˈɡɪvn",
    "forsook": "forsook",
    "forsaken": "fəˈseɪkən",
    "frostbit": "ˈfrɒstbɪt",
    "frostbitten": "ˈfrɒstbɪtn",
    "hewn": "hjuːn",
    "hid": "hɪd",
    "kept": "kept",
    "leant": "lent",
    "leapt": "lept",
    "learnt": "lɜːnt",
    "lie": "laɪ",
    "lay": "leɪ",
    "mown": "məʊn",
    "partaken": "pɑːˈteɪkən",
    "risen": "ˈrɪzn",
    "sawn": "sɔːn",
    "slid": "slɪd",
    "sow": "səʊ",
    "sowed": "səʊd",
    "sown": "səʊn",
    "spelt": "spelt",
    "spilt": "spɪlt",
    "strewn": "struːn",
    "striven": "ˈstrɪvn",
    "sunburnt": "ˈsʌnbɜːnt",
    "taken": "ˈteɪkən",
    "typewrite": "taɪpraɪt",
    "typewrote": "taɪprəʊt",
    "typewritten": "taɪprɪtn",
    # "unbent": "ˌʌnˈbent",
    # "unbind": "ˌʌnbaɪnd",
    # "unbound": "ˌʌnbaʊnd",
    # "unclothe": "ˌʌnkləʊð",
    # "unclothed": "ˌʌnkləʊðd",
    # "unclad": "ˌʌnklæd",
    "waylaid": "weɪˈleɪd",
    "withdraw": "wɪðˈdrɔː",
    "withdrew": "wɪðˈdruː",
    "withdrawn": "wɪðˈdrɔːn",
    "withheld": "wɪðˈheld",
    "withstood": "wɪðˈstʊd"
}

# A table of prefixes' transcriptions obtained by manual observation of the OLD.
PREFIX_TRANSCRIPTIONS = {
    "mis": "ˌmɪsˈ",
    "over": "ˌəʊvəˈ",
    "out": "ˌaʊtˈ",
    "pre": "ˌpriːˈ",
    "re": "ˌriːˈ",
    "un": "ˌʌn",
    "under": "ˌʌndəˈ",
    "de": "diː"
}

# Since OLD is limited in terms of these, the transcriptions below have been fetched from Wiktionary. Might be incorrect.
CONTRACTIONS = {
    "i've": "aɪv",
    "you've": "juːv",
    "we've": "wiːv",
    "they've": "ðeɪv",
    "i'll": "aɪl",
    "you'll": "juːl",
    "he'll": "hɪl",
    "she'll": "ʃɪl",
    "it'll": "ˈɪtl̩",
    "we'll": "wɪl",
    "they'll": "ðeɪl",
    "i'm": "aɪm",
    "you're": "jʊə(ɹ)",
    "we're": "wɪə(ɹ)",
    "they're": "ðɛə(ɹ)",
    "i'd": "aɪd",
    "you'd": "juːd",
    "he'd": "hiːd",
    "she'd": "ʃiːd",
    "it'd": "ˈɪtəd",
    "we'd": "wiːd",
    "they'd": "ðeɪd",
    "it's": "ɪts"
}

class DictionaryParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        # Needed to differentiate between British English and American English transcriptions:
        self.isBritish = False
        self.record = False
        self.found = dict()
        self.counter = 0;
        self.type = ""
        self.headWord = ""
        self.notFound = False
        self.recordError = False

        # Record type - remembers which of the following is being recorded:
        #   t - Type (noun / verb / adjective)
        #   w - Word
        #   h - Head word
        #   e - Error (Not Found)
        self.recordType = ''

    def handle_starttag(self, tag, attrs):
        if attrs is not None and attrs != []:
            if tag == "span":
                if attrs[0][0] == "class":
                    if attrs[0][1] == "phon":
                        if self.isBritish:
                            if self.type == "":
                                self.type = "unknown"
                                self.found[self.type] = []
                            self.record = True
                            self.recordType = 'w'
                            self.isBritish = False
                    if attrs[0][1] == "pos":
                        self.record = True
                        self.recordType = 't'

            elif tag == "div":
                for attr in attrs:
                    if attr[0] == "class":
                        if attr[1] == "phons_br":
                            self.isBritish = True
                    elif attr[0] == "id":
                        if attr[1] == "search-results":
                            self.notFound = True
                    
            elif tag == "h1":
                for attr in attrs:
                    if attr[0] == "class":
                        if attr[1] == "headword":
                            self.record = True
                            self.recordType = 'h'

    def handle_endtag(self, tag):
        if self.record:
            self.record = False
            if self.recordType == 't':
                if self.type == "" or self.type is None:
                    self.type = "unknown"
                self.found[self.type] = []

    def handle_data(self, data):
        if self.record == True:
            if self.recordType == 't':
                self.type += data
            elif self.recordType == 'w':
                self.found[self.type].append(data.strip('/'))
            elif self.recordType == 'h':
                self.headWord = data
            elif self.recordType == 'e':
                if "No exact match found" in data:
                    self.notFound = True

def syllableCount(word):
    count = 0
    if word[0] in VOWELS:
        count += 1
    for index in range(1, len(word)):
        if word[index] in VOWELS and word[index - 1] not in VOWELS:
            count += 1
    if word.endswith("e"):
        count -= 1
    if count == 0:
        count += 1
    return count

def getKeysList(items):
    keysList = list()
    for key in items.keys():
        keysList.append(key)
    return keysList

def getPluralOrThirdPerson(type, word, items):
    multipleItems = False
    lastChar = word[-1]
    lastTwoChars = word[-2:]
    if type == "verb":
        lastTranscriptionChar = items[type][2][-1]
        lastTwoTranscriptionChar = items[type][2][-2:]
        if (lastChar in VOICELESS_CONSONANTS or lastTwoChars in VOICELESS_CONSONANTS) or \
            (lastTranscriptionChar in VOICELESS_CONSONANTS or lastTwoTranscriptionChar in VOICELESS_CONSONANTS):
            items[type][2] += "s"
        elif (lastChar in FRICATIVE_SOUNDS) or (lastTranscriptionChar in FRICATIVE_SOUNDS):
            items[type][2] += "iz"
        elif (lastChar in VOWELS or lastChar in VOICED_CONSONANTS or lastTwoChars in VOICED_CONSONANTS) or \
                (lastTranscriptionChar in VOWELS or lastTranscriptionChar in VOICED_CONSONANTS or lastTwoTranscriptionChar in VOICED_CONSONANTS):
            items[type][2] += "z"
    else:
        pos = 0
        for item in items[type]:
            lastTranscriptionChar = items[type][pos][-1]
            lastTwoTranscriptionChar = items[type][pos][-2:]
            if (lastChar in VOICELESS_CONSONANTS or lastTwoChars in VOICELESS_CONSONANTS) or \
                (lastTranscriptionChar in VOICELESS_CONSONANTS or lastTwoTranscriptionChar in VOICELESS_CONSONANTS):
                items[type][pos] += "s"
            elif (lastChar in FRICATIVE_SOUNDS) or (lastTranscriptionChar in FRICATIVE_SOUNDS):
                items[type][pos] += "iz"
            elif (lastChar in VOWELS or lastChar in VOICED_CONSONANTS or lastTwoChars in VOICED_CONSONANTS) or \
                    (lastTranscriptionChar in VOWELS or lastTranscriptionChar in VOICED_CONSONANTS or lastTwoTranscriptionChar in VOICED_CONSONANTS):
                items[type][pos] += "z"
            pos += 1
            
    return items

def getTranscription(wordToTranscribe, wordType=None):
    data = []
    word = wordToTranscribe

    prefix_re = False
    prefix_un = False
    prefix_out = False
    prefix_mis = False
    prefix_pre = False
    prefix_over = False
    prefix_under = False
    prefix_de = False
    prefix = ""
    prefixExists = False
    prefixIterations = 1
    
    if word[:5] == "under":
        prefix_under = True
        prefixIterations = 2
        prefix = "under"

    elif word[:2] == "re":
        prefix_re = True
        prefixIterations = 2
        prefix = "re"

    elif word[:2] == "un":
        prefix_un = True
        prefixIterations = 2
        prefix = "un"
        
    elif word[:3] == "out":
        prefix_out = True
        prefixIterations = 2
        prefix = "out"

    elif word[:3] == "mis":
        prefix_mis = True
        prefixIterations = 2
        prefix = "mis"

    elif word[:3] == "pre":
        prefix_pre = True
        prefixIterations = 2
        prefix = "pre"

    elif word[:4] == "over":
        prefix_over = True
        prefixIterations = 2
        prefix = "over"   

    elif word[:2] == "de":
        prefix_de = True
        prefixIterations = 2
        prefix = "de" 

    for prefixIteration in range(prefixIterations):
        # If this is the second iteration, this means that the original word has a prefix and a match for the prefixed word
        # hasn't been found - therefore, remove the prefix and try adding it manually later.
        # This is needed in some cases where it might seem like there is a prefix but there actually isn't one,
        # for example "outside" 
        if prefixIteration == 1:
            # word = prefix + word
            if prefix_re or prefix_un or prefix_de:
                word = word[2:]

            elif prefix_out or prefix_mis or prefix_pre:
                word = word[3:]

            elif prefix_over:
                word = word[4:]

            elif prefix_under:
                word = word[5:]
        
        # First, check if the word is present in the list of irregular verbs and, if so, return it.
        for irregular_verb in IRREGULAR_VERBS:
            if irregular_verb == word:
                transcription = dict()
                transcription["verb"] = list()
                # Because the part of the script that takes out the results looks at the second item in the list, there should be a placeholder.
                transcription["verb"].append(None)
                
                if prefixIteration == 0:
                    transcription["verb"].append(IRREGULAR_VERBS[irregular_verb])
                else:
                    transcription["verb"].append(PREFIX_TRANSCRIPTIONS[prefix] + IRREGULAR_VERBS[irregular_verb])
                
                transcription["verb"].append(IRREGULAR_VERBS[irregular_verb])
                data.append(transcription)
                return data

        # If not, check if it's a contraction and get the actual determiner - the verb is added later.
        for contraction in CONTRACTIONS:
            if word == contraction:
                returnDict = dict()
                returnDict["determiner, contraction"] = list()
                returnDict["determiner, contraction"].append(CONTRACTIONS[contraction])
                data.append(returnDict)
                return data

        # Start querying the Oxford Learner's Dictionaries.

        isPossessive = False
        isPluralOrThirdPerson = False
        endsInLy = False
        endsInEr = False
        addTranscriptionLy = False
        addTranscriptionEr = False

        if word[-2:] == "'s" or word[-2:] == "’s":
            word = word[:-2]
            isPossessive = True
        elif (word[-4:] == "sses") or (word[-1] == "s" and word[-4:] != "ness"):
            isPluralOrThirdPerson = True
        if word[-2:] == "ly":
            endsInLy = True
        if word[-2:] == "er":
            endsInEr = True
        
        iterateURLs = True
        iteration = 1

        try:
            http = urllib3.PoolManager()

            # If the word doesn't have any homonyms, only look through the content on this page.
            # Otherwise, keep iterating until the server returns "Not Found".
            while iterateURLs:
                URL = baseURL + word.strip()
                if VERBOSE_DEBUG: 
                    print ("---- NEW ITERATION -", iteration, "----")
                    print ("[*] [DEBUG] Word:", word)
                parser = DictionaryParser()
                openURL = http.request("GET", URL)
                if VERBOSE_DEBUG: print ("[*] [DEBUG] Opening URL:", URL)
                redirectedURL = openURL.geturl()
                if VERBOSE_DEBUG: print ("[*] [DEBUG] Redirecting to URL:", redirectedURL)

                if openURL.status == 200:
                    if VERBOSE_DEBUG: print ("[*] [DEBUG] Got status 200")
                    parser.feed(openURL.data.decode())
                    iterateURLs = False
                    if parser.notFound:
                        if word[-1] == 's':
                            # isPluralOrThirdPerson = True
                            word = word[:-1]
                            iterateURLs = True
                        elif endsInLy:
                            word = word[:-2]
                            iterateURLs = True
                            endsInLy = False
                            addTranscriptionLy = True
                        elif endsInEr:
                            word = word[:-2]
                            iterateURLs = True
                            endsInEr = False
                            addTranscriptionEr = True
                elif openURL.status == 404:
                    if VERBOSE_DEBUG: print ("[*] [DEBUG] Got status 404")
                    if VERBOSE_DEBUG: print ("[*] [DEBUG] URL", URL, "returned HTTP 404.")
                    iterateURLs = False
                else:
                    if VERBOSE_DEBUG: print ("[*] [DEBUG] Got status", openURL.status, "- Stopping iterations")
                    else: print("An error occurred.")
                    iterateURLs = False

                if bool(parser.found) != False:
                    items = parser.found
                    
                    if addTranscriptionEr:
                        print ("Got items:", items)
                        changeItems = getKeysList(items)
                        for i in range(len(changeItems[0])):
                            if items[changeItems[0]][i][-3:] == "(r)":
                                items[changeItems[0]][i] = items[changeItems[0]][i][:-3] + "r"
                            items[changeItems[0]][i] += "ə(r)"
                        print ("Items out:", items)
                    
                    if "noun" in items:
                        if iteration == 1 and isPluralOrThirdPerson and parser.headWord == word:
                            data.append(items)
                        elif isPossessive or (word[-1] == 's' and word != parser.headWord) or isPluralOrThirdPerson:
                            data.append(getPluralOrThirdPerson("noun", parser.headWord, items))
                            if VERBOSE_DEBUG: print ("[*] [DEBUG] Word is possessive or plural, getting transcription...")
                        else:
                            data.append(items)
                    elif "adjective" in items:
                        if isPossessive or (word[-1] == 's' and word != parser.headWord) or isPluralOrThirdPerson:
                            data.append(getPluralOrThirdPerson("adjective", parser.headWord, items))
                            if VERBOSE_DEBUG: print ("[*] [DEBUG] Word is possessive or plural, getting transcription...")

                        # https://youglish.com/pronounce/busiest/english offers a different transcription for -er and -est.
                        elif parser.headWord[:-1] in word and parser.headWord != word:
                            if word[-2:] == "er":
                                for key in items:
                                    items["adjective"][0] += "ə"
                            elif word[-3:] == "est":
                                for item in items["adjective"]:
                                    items["adjective"][0] += "ɪst"
                            data.append(items)
                        elif addTranscriptionLy:
                            for i in range(len(items["adjective"])):
                                items["adjective"][i] += "lɪ"
                            data.append(items)
                        else:
                            data.append(items)
                    elif "verb" in items:
                        data.append(items)
                    else:
                        if isPossessive or (word[-1] == 's' and word != parser.headWord) or isPluralOrThirdPerson:
                            data.append(getPluralOrThirdPerson(getKeysList(items)[0], parser.headWord, items))
                            if VERBOSE_DEBUG: print ("[*] [DEBUG] Word is possessive or plural, getting transcription...")
                        elif isPluralOrThirdPerson:
                            data.append(getPluralOrThirdPerson(getKeysList(items)[0], parser.headWord, items))
                        else:
                            data.append(items)
                
                if VERBOSE_DEBUG: 
                    print ("[*] DEBUG Data so far:", data)
                    print ("---- END ITERATION -", iteration, "----")
                iteration += 1
                parser.close()
        except Exception as e:
            print ("Something went wrong, check the following details for more information:")
            print ("Exception Type:", type(e))
            print ("Arguments", e.args)
            print ("Exception txt:", e)

        if prefixIteration == 1:
            for dictionary in data:
                for key in list(dictionary):
                    arrLocation = 0
                    for transcription in dictionary[key]:
                        dictionary[key][arrLocation] = PREFIX_TRANSCRIPTIONS[prefix] + transcription
                        arrLocation += 1

        if (len(data) != 0):
            if VERBOSE_DEBUG: print ("[*] [DEBUG] Returning", data)
            return data
        else:
            if VERBOSE_DEBUG: print ("[*] [DEBUG] Got no data to return")

# If this is a complex word, e.g. inter-change, split the words, get transcriptions for each and merge the results
def getComplexTranscription(wordsCombination):
    if "year" not in wordsCombination:
        data = getTranscription(wordsCombination)
        if (data != None):
            for dictionary in data:
                for key in dictionary:
                    for element in dictionary[key]:
                        return element

    words = split(" |-|—|/|\[|\]", wordsCombination)
    endResult = dict()
    tempArray = ""
    for word in words:
        data = getTranscription(word)
        for dictionary in data:
            for key in dictionary:
                if key == "verb":
                    if word[-3:] == "ing":
                        if syllableCount(word) == 1:
                            tempArray += dictionary["verb"][0]
                        else:
                            if dictionary["verb"][-1][-2:] == "ɪŋ":
                                tempArray += dictionary["verb"][-1] + " "
                            elif dictionary["verb"][-2][-2:] == "ɪŋ":
                                tempArray += dictionary["verb"][-2] + " "
                            else:
                                for transcription in dictionary["verb"]:
                                    if transcription[-2:] == "ɪŋ":
                                        tempArray += transcription + " "
                    elif word[-1] == "s":
                        if word[-2:] == "ss":
                            tempArray += dictionary["verb"][1] + " "
                        else:
                            tempArray += dictionary["verb"][2] + " "
                    elif word[-2:] == "ed":
                        tempArray += dictionary["verb"][3] + " "
                    else:
                        tempArray += dictionary["verb"][1] + " "
                else:
                    tempArray += dictionary[key][0] + " "
                    break
    return tempArray.rstrip()

def updateProgress(thisWord, totalWords, word, errorCount):
    percentage = ("%.2f" % ((thisWord / totalWords) * 100))
    sys.stdout.write("\rProgress: {0}/{1} ({2}%) ||| Errors: {3} ||| Current word is: {4}".format(thisWord, totalWords, percentage, errorCount, word))
    sys.stdout.flush()
    sys.stdout.write("\033[K")
    sys.stdout.flush()
    if (thisWord == totalWords):
        print ("\r\n")
    return

# Check if the file provided via argument is a valid file.
def is_valid_file(parser, arg):
    if not os.path.exists(arg):
        parser.error("The file %s does not exist!" % arg)
        sys.exit()
    return arg

argParser = argparse.ArgumentParser(description="Get transcriptions for words from the Oxford Learner's Dictionaries.", allow_abbrev=False)
argParser.add_argument("-v", "--verbose", help="Produce additional DEBUG output.", action="store_true")
argParser.add_argument("-t", "--plaintext", help="If the file that has to be transcribed is not in Excel, use this flag.", action="store_true")
argParser.add_argument('-f', "--file", help='Path to a .xls(x) or .txt file containing words to be transcribed. If .txt file, use -t flag', type=lambda x: is_valid_file(argParser, x), nargs=1, metavar='[File to get word forms from]')
argParser.add_argument("-o", "--output", help="Save output to a separate file.", action="store", type=argparse.FileType('w'), nargs=1)

args = argParser.parse_args()

VERBOSE_DEBUG = args.verbose
PLAINTEXT = args.plaintext

if args.file is not None:
    FILENAME = args.file[0]
else:
    FILENAME = "transcribe.xlsx"

if FILENAME[-3:] == "txt" and not PLAINTEXT:
    print ("It looks like you're pointing to a .txt file - please use -t.")
    sys.exit()
elif FILENAME[-3:] == "xls" or FILENAME[-4:] == "xlsx" and PLAINTEXT:
    print ("Looks like you're pointing to a .xls(x) file, please omit -t.")
    sys.exit()

if PLAINTEXT:
    with open(FILENAME, 'r') as file:
        for word in file:
            try:
                # Ignore commented words
                if word[0] != '#':
                    word = word.strip(CHARACTERS_TO_STRIP)
                    word = word.replace("’", "'")
                    if word != '':
                        shouldContinue = True
                        complex = False
                        for separator in WORD_SEPARATORS:
                            if separator in word:
                                result = getComplexTranscription(word)
                                print (word, "(complex)")
                                print ("\t" + result)
                                shouldContinue = False
                        if shouldContinue:
                            word = word.lower()
                            result = getTranscription(word)
                            for transcriptions in result:
                                for wordType in transcriptions:
                                    print(word + ' (' + wordType + ')')
                                    if wordType == "verb":
                                        if word[-3:] == "ing":
                                            if syllableCount(word) == 1:
                                                print ("\t" + transcriptions["verb"][0])
                                            else:
                                                if transcriptions["verb"][-1][-2:] == "ɪŋ":
                                                    print ("\t" + transcriptions["verb"][-1])
                                                elif transcriptions["verb"][-2][-2:] == "ɪŋ":
                                                    print ("\t" + transcriptions["verb"][-2])
                                                else:
                                                    for transcription in transcriptions["verb"]:
                                                        if transcription[-2:] == "ɪŋ":
                                                            print ("\t" + transcription)
                                        elif word[-1] == "s":
                                            if word[-2:] == "ss":
                                                print ("\t" + transcriptions["verb"][1])
                                            else:
                                                print ("\t" + transcriptions["verb"][2])
                                        elif word[-2:] == "ed":
                                            print ("\t" + transcriptions["verb"][3])
                                        else:
                                            print ("\t" + transcriptions["verb"][1])
                                    else:
                                        for transcription in transcriptions[wordType]:
                                            print("\t" + transcription)
            except Exception as e:
                print ("An error occurred at word \"{0}\". Message: {1}".format(word, e))
else:
    workbook = load_workbook(filename=FILENAME)
    sheet = workbook.active
    totalWords = sheet.max_row
    i = 1
    errorCount = 0
    rPos = "A" + str(1)
    wPos = "B" + str(1)
    while sheet[rPos].value != None:
        if sheet[wPos].value == None:
            shouldContinue = True
            word = sheet[rPos].value
            word = word.strip(CHARACTERS_TO_STRIP)
            word = word.replace("’", "'")
            updateProgress(i, totalWords, word, errorCount)
            complex = False
            try:
                for separator in WORD_SEPARATORS:
                    if separator in word:
                        result = getComplexTranscription(word)
                        sheet[wPos] = result
                        wPos = str(chr(ord(wPos[0]) + 1)) + str(i)
                        shouldContinue = False
                if shouldContinue:
                    word = word.lower()
                    result = getTranscription(word)
                    for transcriptions in result:
                        for wordType in transcriptions:
                            transcribed = ""
                            if wordType == "verb":
                                if word[-3:] == "ing":
                                    if syllableCount(word) == 1:
                                        transcribed = transcriptions["verb"][0]
                                    else:
                                        if transcriptions["verb"][-1][-2:] == "ɪŋ":
                                            transcribed = transcriptions["verb"][-1]
                                        elif transcriptions["verb"][-2][-2:] == "ɪŋ":
                                            transcribed = transcriptions["verb"][-2]
                                        else:
                                            for transcription in transcriptions["verb"]:
                                                if transcription[-2:] == "ɪŋ":
                                                    transcribed = transcription
                                    
                                elif word[-1] == "s":
                                    if word[-2:] == "ss":
                                        transcribed = transcriptions["verb"][1]
                                    else:
                                        transcribed = transcriptions["verb"][2]
                                elif word[-2:] == "ed":
                                    transcribed = transcriptions["verb"][3]
                                else:
                                    transcribed = transcriptions["verb"][1]
                            else:
                                for transcription in transcriptions[wordType]:
                                    if transcription != "":
                                        transcribed = "\r\n" + transcription
                            sheet[wPos] = transcribed
                            wPos = str(chr(ord(wPos[0]) + 1)) + str(i)
            except Exception as e:
                errorCount += 1
        if (i % 50 == 0):
            workbook.save(filename=FILENAME)
        i += 1
        rPos = "A" + str(i)
        wPos = "B" + str(i)

    workbook.save(filename=FILENAME)
    workbook.close()
    if (errorCount > 0):
        print ("\n\r\n\rFinished with {0} error(s). Check the file you supplied (transcribe.xlsx by default).".format(errorCount))
    else:
        print ("\n\r\n\rAll done - no problems encountered. Check the file you supplied (transcribe.xlsx by default).")